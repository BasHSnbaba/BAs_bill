import openpyxl
from datetime import datetime
from PySide2.QtGui import QIcon
from PySide2.QtWidgets import QApplication, QMessageBox, QTableWidgetItem
from PySide2.QtUiTools import QUiLoader
from PySide2.QtCore import QTimer

class Stats:
    def __init__(self):

        # 载入ui文件
        self.ui = QUiLoader().load('BAs_bill/ui/main.ui')

        # 设置按钮 showSetting
        self.ui.pushButton_setting.clicked.connect(self.showSetting)

        # 显示时间 showTime
        self.ui.nowTime.setText(str(datetime.now()).split('.')[0])
        self.timer = QTimer()
        self.timer.timeout.connect(self.showTime)
        self.timer.start(1000)

        # 时间项目 timeOSadd
        self.timeOScnt = 0  # 时间项目ID
        self.ui.pushButton_timeIDadd.clicked.connect(self.timeOSadd)

        # 时间项目 timeOSstop pushButton_timeIDstop
        self.ui.pushButton_timeIDstop.clicked.connect(self.timeOSstop)

        # 时间项目 timeOSdel pushButton_timeIDdel
        self.ui.pushButton_timeIDdel.clicked.connect(self.timeOSdel)

        # 商品添加系统
        self.ui.pushButton_addItem.clicked.connect(self.addItem)

        # 账单导出按钮
        self.ui.pushButton_ex.clicked.connect(self.ex)

    def ex(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.active
        self.sheet.title = str(datetime.now()).split(' ')[0] + '账单'

        self.sheet['A1'] = '项目id'
        self.sheet['B1'] = '项目类别'
        self.sheet['C1'] = '项目实体id'
        self.sheet['D1'] = '开始时间'
        self.sheet['E1'] = '结束时间'
        self.sheet['F1'] = '金额'

        # 表项目循环
        for i in range(self.timeOScnt):
            for j in range(6):
                if j == 5:
                    self.sheet.cell(row=i+2,column=j+1).value = float(self.ui.tableWidget.item(i,j).text())
                else:
                    self.sheet.cell(row=i+2,column=j+1).value = self.ui.tableWidget.item(i,j).text()
                
        # 表价格汇总
        self.sheet['E'+str(i+3)] = '金额汇总:'
        self.sum = 0.0
        for k in range(self.timeOScnt):
            self.sum = self.sum+self.sheet.cell(row=k+2,column=6).value
        self.sheet['F'+str(i+3)] = self.sum

        self.wb.save(str(datetime.now()).split(' ')[0] + '账单.xlsx')


    def addItem(self):
        if self.ui.lineEdit_p.text() != '':
            self.method = self.ui.comboBox_c.currentText()
            self.newItemID = QTableWidgetItem(str(self.timeOScnt))
            self.newItemClass = QTableWidgetItem(str(self.method))
            self.newItemEntity = QTableWidgetItem('None')
            self.newItemStartDate = QTableWidgetItem('None')
            self.newItemStopDate = QTableWidgetItem('None')
            self.newItemPrice = QTableWidgetItem(str(self.ui.lineEdit_p.text()))
            self.ui.tableWidget.setItem(self.timeOScnt,0,self.newItemID)
            self.ui.tableWidget.setItem(self.timeOScnt,1,self.newItemClass)
            self.ui.tableWidget.setItem(self.timeOScnt,2,self.newItemEntity)
            self.ui.tableWidget.setItem(self.timeOScnt,3,self.newItemStartDate)
            self.ui.tableWidget.setItem(self.timeOScnt,4,self.newItemStopDate)
            self.ui.tableWidget.setItem(self.timeOScnt,5,self.newItemPrice)
            self.timeOScnt = self.timeOScnt + 1
            self.ui.lineEdit_p.clear()
        else:
            QMessageBox.warning(
            self.ui,
            '错误',
            f'''请先输入商品价格!'''
            )

    def showTime(self):
        self.ui.nowTime.setText(str(datetime.now()).split('.')[0])

    def showSetting(self):
        QMessageBox.information(
            self.ui,
            '设置',
            f'''Author: BasHSnbaba
Version: bill v0.01'''
        )

    def timeOSadd(self):
        if self.ui.lineEdit_timeID.text() != '':
            self.newItemID = QTableWidgetItem(str(self.timeOScnt))
            self.newItemClass = QTableWidgetItem('台球')
            self.newItemEntity = QTableWidgetItem(str(self.ui.lineEdit_timeID.text()))
            self.newItemStartDate = QTableWidgetItem(str(datetime.now()).split('.')[0].split(' ')[1])
            self.ui.tableWidget.setItem(self.timeOScnt,0,self.newItemID)
            self.ui.tableWidget.setItem(self.timeOScnt,1,self.newItemClass)
            self.ui.tableWidget.setItem(self.timeOScnt,2,self.newItemEntity)
            self.ui.tableWidget.setItem(self.timeOScnt,3,self.newItemStartDate)
            self.timeOScnt = self.timeOScnt + 1
            self.ui.lineEdit_timeID.clear()
        else:
            QMessageBox.warning(
            self.ui,
            '错误',
            f'''请先输入实体项目编号!'''
            )
    
    def timeOSstop(self):
        self.currentItem = self.ui.tableWidget.currentRow() # 选中的行数
        self.newItemStopDate = QTableWidgetItem(str(datetime.now()).split('.')[0].split(' ')[1]) # 停止时间
        if self.currentItem == -1:
            QMessageBox.warning(
            self.ui,
            '错误',
            f'''请先在表格中选择一个项目!'''
            )
        else:
            if (self.currentItem+1 <= self.timeOScnt) and (self.timeOScnt != 0):
                self.replay = QMessageBox.question(
                self.ui,
                '提示',
                '您确定要停止 项目id为:'+str(self.currentItem)+' 的项目吗?',
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if self.replay == QMessageBox.Yes:
                    self.ui.tableWidget.setItem(self.currentItem,4,self.newItemStopDate)
                    self.timeStop = self.newItemStopDate.text().split(':')
                    self.timeStart = self.newItemStartDate.text().split(':')
                    self.time = (int(self.timeStop[0])-int(self.timeStart[0]))*60+int(self.timeStop[1])-int(self.timeStart[1])

                    if self.time <= 60:
                        self.money = 10
                    else:
                        self.money = int(self.time*0.17)
                    
                    self.newItemMoney = QTableWidgetItem(str(self.money))
                    self.ui.tableWidget.setItem(self.currentItem,5,self.newItemMoney)
                else:
                    QMessageBox.about(
                    self.ui,
                    '取消信息',
                    f'''您已取消操作.'''
                    )
            else:
                QMessageBox.warning(
                self.ui,
                '错误',
                f'''您选择的单元格没有项目!'''
                )

    def timeOSdel(self):
        self.currentItem = self.ui.tableWidget.currentRow() # 选中的行数
        if self.currentItem == -1:
            QMessageBox.warning(
            self.ui,
            '错误',
            f'''请先在表格中选择一个项目!'''
            )
        else:
            if (self.currentItem+1 <= self.timeOScnt) and (self.timeOScnt != 0):
                self.replay = QMessageBox.question(
                self.ui,
                '提示',
                '您确定要删除 项目id为:'+str(self.currentItem)+' 的项目吗?',
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
                )
                if self.replay == QMessageBox.Yes:
                    self.newItemDelText = QTableWidgetItem('已删除项目')
                    self.ui.tableWidget.setItem(self.currentItem,4,self.newItemDelText)
                    
                    self.newItemDelMoney = QTableWidgetItem('0')
                    self.ui.tableWidget.setItem(self.currentItem,5,self.newItemDelMoney)
                else:
                    QMessageBox.about(
                    self.ui,
                    '取消信息',
                    f'''您已取消操作.'''
                    )
            else:
                QMessageBox.warning(
                self.ui,
                '错误',
                f'''您选择的单元格没有项目!'''
                )

app = QApplication([])
app.setWindowIcon(QIcon('bill.png'))
stats = Stats()
stats.ui.show()
app.exec_()